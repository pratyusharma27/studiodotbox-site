"""
Studio Dotbox - generate_pages.py

Master content generator. Run this whenever you update studiodotbox-content.xlsx
or whenever you want to rebuild all the static pages.

Usage:
    python3 generate_pages.py
"""
import os
import sys
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl")
    print("Install it with: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent.resolve()
EXCEL_FILE = SCRIPT_DIR / 'studiodotbox-content.xlsx'

sys.path.insert(0, str(SCRIPT_DIR))
sys.path.insert(0, '/home/claude')

# We will build everything in this folder
SITE_ROOT = SCRIPT_DIR

# Web3Forms access key is injected at build time from the environment so the
# real key never lives in committed source. Falls back to the placeholder,
# which the pre-flight audit treats as a hard failure.
WEB3FORMS_KEY = os.environ.get('WEB3FORMS_KEY', 'YOUR_WEB3FORMS_KEY_HERE')

# Google Analytics 4 measurement ID (format: G-XXXXXXXXXX). Injected into the
# <head> of every page via write_page(). A measurement ID is not a secret (it
# ships in public page source), so it is safe to hardcode the default here;
# it can still be overridden per-build with a GA_MEASUREMENT_ID env var on
# Netlify. When the value is left as the placeholder, no analytics tag is
# emitted, so builds never ship a broken or fake tag.
GA_MEASUREMENT_ID = os.environ.get('GA_MEASUREMENT_ID', 'G-WMKRFRGXS8')

# Canonical host for this site. www is the primary domain; canonical tags,
# sitemap URLs, and robots all point here.
BASE_URL = "https://www.studiodotbox.com"

# Populated by write_page as pages are built, then consumed by build_sitemap.
PAGES_BUILT = []


def canonical_url(filename):
    """Map an output filename to its canonical www https URL (clean, no .html)."""
    path = filename.replace(os.sep, "/")
    if path == "index.html":
        return BASE_URL + "/"
    if path.endswith("/index.html"):
        path = path[:-len("index.html")]
    elif path.endswith(".html"):
        path = path[:-len(".html")]
    return BASE_URL + "/" + path


# ==================== TEMPLATES ====================

# Google Analytics 4 (gtag.js). Kept out of HEAD_TEMPLATE's .format() call so
# the JavaScript braces below are never treated as format fields. The literal
# __GA_ID__ token is swapped for the real measurement ID in write_page().
GA_SNIPPET_TEMPLATE = """  <!-- Google tag (gtag.js) -->
  <script async src="https://www.googletagmanager.com/gtag/js?id=__GA_ID__"></script>
  <script>
    window.dataLayer = window.dataLayer || [];
    function gtag(){dataLayer.push(arguments);}
    gtag('js', new Date());
    gtag('config', '__GA_ID__');
  </script>
"""

HEAD_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
{ga}  <title>{title}</title>
  <meta name="description" content="{description}">
  {canonical}

  <link rel="icon" type="image/svg+xml" href="{css_prefix}images/favicon.svg">
  <link rel="alternate icon" type="image/png" href="{css_prefix}images/favicon.png">
  <link rel="apple-touch-icon" href="{css_prefix}images/apple-touch-icon.png">

  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;500;700&display=swap" rel="stylesheet">

  <link rel="stylesheet" href="{css_prefix}css/main.css">
  <link rel="stylesheet" href="{css_prefix}css/pages.css">
  {extra_css}
</head>
<body{body_class}>
"""

HEADER_TEMPLATE = """
  <!-- ===== HEADER ===== -->
  <header class="site-header">
    <div class="site-header__inner">
      <a href="{home}" class="site-header__logo" aria-label="Studio Dotbox, home">
        <img src="{img_prefix}images/logo.svg" alt="Studio Dotbox">
      </a>
      <nav class="site-header__nav" aria-label="Primary">
        <a href="{prefix}work.html">Work</a>
        <a href="{prefix}studio.html">Studio</a>
        <a href="{prefix}services.html">Services</a>
        <a href="{prefix}journal.html">Journal</a>
        <a href="{prefix}contact.html">Contact</a>
      </nav>
      <button class="site-header__menu-toggle" aria-label="Open menu" aria-expanded="false">
        <span></span><span></span><span></span>
      </button>
    </div>
  </header>

  <div class="mobile-menu" aria-hidden="true">
    <button class="mobile-menu__close" aria-label="Close menu">×</button>
    <div class="mobile-menu__primary">
      <a href="{prefix}work.html">Work</a>
      <a href="{prefix}studio.html">Studio</a>
      <a href="{prefix}services.html">Services</a>
      <a href="{prefix}journal.html">Journal</a>
      <a href="{prefix}contact.html">Contact</a>
    </div>
    <div class="mobile-menu__divider"></div>
    <div class="mobile-menu__secondary">
      <a href="{prefix}philosophy.html">Philosophy</a>
      <a href="{prefix}how-we-work.html">How we work</a>
    </div>
    <div class="mobile-menu__divider"></div>
    <div class="mobile-menu__secondary">
      <a href="https://www.linkedin.com/company/studio-dotbox/" target="_blank" rel="noopener">LinkedIn</a>
      <a href="https://www.instagram.com/studiodotbox/" target="_blank" rel="noopener">Instagram</a>
    </div>
  </div>
  <div class="mobile-menu__overlay"></div>
"""

FOOTER_TEMPLATE = """
  <!-- ===== FOOTER ===== -->
  <footer class="site-footer">
    <div class="site-footer__inner">
      <div class="site-footer__columns">

        <div class="site-footer__column">
          <a href="{home}" class="site-footer__logo" aria-label="Studio Dotbox, home">
            <img src="{img_prefix}images/logo.svg" alt="Studio Dotbox">
          </a>
          <p>Architecture, interiors, and urban planning.</p>
          <p>A studio-based practice, based in Noida.</p>
        </div>

        <div class="site-footer__column">
          <h4>Contact</h4>
          <a href="mailto:studiodotbox@gmail.com">studiodotbox@gmail.com</a>
          <a href="tel:+919571129136">+91 9571 129 136</a>
        </div>

        <div class="site-footer__column">
          <h4>Navigate</h4>
          <a href="{home}">Home</a>
          <a href="{prefix}work.html">Work</a>
          <a href="{prefix}studio.html">Studio</a>
          <a href="{prefix}services.html">Services</a>
          <a href="{prefix}philosophy.html">Philosophy</a>
          <a href="{prefix}how-we-work.html">How we work</a>
          <a href="{prefix}journal.html">Journal</a>
          <a href="{prefix}contact.html">Contact</a>
        </div>

        <div class="site-footer__column">
          <h4>Follow</h4>
          <a href="https://www.linkedin.com/company/studio-dotbox/" target="_blank" rel="noopener">LinkedIn</a>
          <a href="https://www.instagram.com/studiodotbox/" target="_blank" rel="noopener">Instagram</a>

          <div class="site-footer__subscribe">
            <p>New writing from the studio, once or twice a month.</p>
            <form id="subscribe-form" action="https://api.web3forms.com/submit" method="POST">
              <input type="hidden" name="access_key" value="YOUR_WEB3FORMS_KEY_HERE">
              <input type="hidden" name="subject" value="Studio Dotbox Journal subscriber">
              <input type="email" name="email" placeholder="Your email" required>
              <button type="submit">Subscribe</button>
            </form>
            <p id="subscribe-success" style="display:none; color: var(--color-text-muted);">Thank you. New writing will land in your inbox.</p>
          </div>
        </div>

        <div class="site-footer__column">
          <h4>Work with us</h4>
          <a href="{prefix}work-with-us.html#professionals">For professionals →</a>
          <a href="{prefix}work-with-us.html#contractors">For contractors →</a>
          <a href="{prefix}work-with-us.html#vendors">For vendors →</a>
        </div>

      </div>

      <div class="site-footer__bar">
        <span>© Studio Dotbox, 2026.</span>
        <div>
          <a href="{prefix}terms.html">Terms</a>
          <a href="{prefix}privacy.html">Privacy</a>
        </div>
      </div>
    </div>
  </footer>

  <script src="{img_prefix}js/main.js"></script>
  <script>
    (function() {{
      var sf = document.getElementById('subscribe-form');
      if (!sf) return;
      var ok = document.getElementById('subscribe-success');
      sf.addEventListener('submit', function(e) {{
        e.preventDefault();
        var data = new FormData(sf);
        var ak = data.get('access_key') || '';
        if (!ak || ak.indexOf('YOUR_') === 0) {{
          alert('Subscribe form will work once the Web3Forms key is added.');
          return;
        }}
        fetch('https://api.web3forms.com/submit', {{ method: 'POST', body: data }})
          .then(function(r) {{ return r.json(); }})
          .then(function(j) {{
            if (j.success) {{ sf.style.display = 'none'; if (ok) ok.style.display = 'block'; }}
            else {{ alert('Sorry, something went wrong. Please try again.'); }}
          }})
          .catch(function() {{ alert('Sorry, something went wrong. Please try again.'); }});
      }});
    }})();
  </script>
  {extra_js}
</body>
</html>
"""


def write_page(filename, title, description, body, depth=0,
               with_animations=False, intro_overlay=None, body_class="",
               extra_stylesheets=None):
    prefix = "../" * depth if depth else ""
    home = "../index.html" if depth else "index.html"

    extra_css = ""
    extra_js = ""
    if with_animations:
        extra_css = f'<link rel="stylesheet" href="{prefix}css/animations.css">'
        extra_js = f'<script src="{prefix}js/animations.js"></script>'
    if extra_stylesheets:
        for sheet in extra_stylesheets:
            extra_css += f'\n  <link rel="stylesheet" href="{prefix}{sheet}">'

    body_class_attr = f' class="{body_class}"' if body_class else ""

    # Canonical tag (www https). Skip for the 404 page.
    if filename == "404.html":
        canonical_tag = ""
    else:
        canonical_tag = f'<link rel="canonical" href="{canonical_url(filename)}">'
        PAGES_BUILT.append(filename)

    # Google Analytics: emit the gtag.js block only when a real measurement ID
    # is configured. While GA_MEASUREMENT_ID is left as the placeholder, ga_tag
    # is empty so no broken/fake tag is shipped.
    if GA_MEASUREMENT_ID and GA_MEASUREMENT_ID != 'G-XXXXXXXXXX':
        ga_tag = GA_SNIPPET_TEMPLATE.replace('__GA_ID__', GA_MEASUREMENT_ID)
    else:
        ga_tag = ""

    html = HEAD_TEMPLATE.format(
        title=title, description=description, canonical=canonical_tag,
        css_prefix=prefix, extra_css=extra_css, body_class=body_class_attr,
        ga=ga_tag
    )

    if intro_overlay:
        html += intro_overlay

    html += HEADER_TEMPLATE.format(prefix=prefix, img_prefix=prefix, home=home)
    html += body
    html += FOOTER_TEMPLATE.format(prefix=prefix, img_prefix=prefix, home=home, extra_js=extra_js)

    # Inject the Web3Forms access key into hidden form inputs only (not the
    # JS guard, which keys off the 'YOUR_' prefix independently).
    html = html.replace('value="YOUR_WEB3FORMS_KEY_HERE"', f'value="{WEB3FORMS_KEY}"')

    output_path = SITE_ROOT / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding='utf-8')
    print(f"Built: {filename}")


# ==================== EXCEL READING ====================

def normalise_text(text):
    """Convert apostrophes to typographic quotes for HTML."""
    if text is None:
        return ""
    text = str(text)
    text = text.replace("'", "&rsquo;").replace('"', '&ldquo;')
    return text

def paragraphs_to_html(text):
    """Convert text with double-line-break paragraphs to HTML p tags."""
    if not text:
        return ""
    text = str(text).strip()
    paras = re.split(r'\n\s*\n', text)
    return "\n".join(f"          <p>{p.strip()}</p>" for p in paras if p.strip())


def count_gallery_images(slug):
    """Count the actual gallery images in a project's image folder.

    The gallery renders sequential files 01.jpg, 02.jpg, ... (hero.jpg and
    thumb.jpg are excluded by site convention). This counts the contiguous
    run starting at 01.jpg, which matches exactly how the gallery is built,
    so the page never renders a placeholder for a missing file. This is the
    source of truth for the image count; Excel Column I is ignored.
    """
    folder = SCRIPT_DIR / 'images' / 'projects' / slug
    if not folder.is_dir():
        return 0
    n = 0
    while (folder / f"{n + 1:02d}.jpg").exists():
        n += 1
    return n


def read_excel():
    """Read the content workbook. Returns (projects, articles)."""
    if not EXCEL_FILE.exists():
        print(f"Warning: {EXCEL_FILE} not found. Skipping content from Excel.")
        return [], []

    wb = load_workbook(EXCEL_FILE, data_only=True)

    projects = []
    if 'Projects' in wb.sheetnames:
        sheet = wb['Projects']
        # Row 1 is header, row 2 is hint, row 3+ is data
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            slug = str(row[0]).strip().lower().replace(' ', '-')
            project = {
                'slug': slug,
                'name': row[1] or '[Project name]',
                'sector': normalise_sector(row[2]),
                'stage': str(row[3] or 'completed').strip().lower().replace(' ', '-'),
                'year': row[4] or '',
                'location': row[5] or '',
                'description': row[6] or '',
                'narrative': row[7] or '',
                'gallery_count': count_gallery_images(slug),
                'materials': row[9] or '',
                'floor_area': row[10] or '',
                'period': row[11] or '',
                'collaborators': row[12] or '',
                'featured': str(row[13] or '').strip().upper() == 'YES',
            }
            projects.append(project)

    articles = []
    if 'Journal' in wb.sheetnames:
        sheet = wb['Journal']
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            slug = str(row[0]).strip().lower().replace(' ', '-')
            # Format publish date if it's a datetime
            date_value = row[4]
            if hasattr(date_value, 'strftime'):
                date_str = date_value.strftime('%d %B %Y')
            else:
                date_str = str(date_value or '')

            article = {
                'slug': slug,
                'title': row[1] or '[Article title]',
                'category': str(row[2] or '').strip(),
                'reading_time': row[3] or '',
                'publish_date': date_str,
                'excerpt': row[5] or '',
                'body': row[6] or '',
                'featured': str(row[7] or '').strip().upper() == 'YES',
            }
            articles.append(article)

    return projects, articles


def category_label(slug):
    """Convert category slug like 'design-thinking' into 'Design thinking'."""
    if not slug:
        return ''
    return slug.replace('-', ' ').strip().capitalize()


def normalise_sector(raw):
    """Map Excel sector values to the site's seven-sector taxonomy.

    Site sectors: residential, retail, commercial, hospitality,
    urban-planning, consultancy, institutional.

    Excel may contain free-form values; this maps them to the canonical slug.
    """
    if not raw:
        return 'residential'
    s = str(raw).strip().lower()
    # Direct matches
    direct = {
        'residential': 'residential',
        'retail': 'retail',
        'commercial': 'commercial',
        'hospitality': 'hospitality',
        'urban planning': 'urban-planning',
        'urban-planning': 'urban-planning',
        'consultancy': 'consultancy',
        'institutional': 'institutional',
    }
    if s in direct:
        return direct[s]
    # Folding decided with Pratyush: Interior Design -> Commercial,
    # Furniture/Exhibition/Event Design -> Retail
    if s in ('interior design', 'interior-design'):
        return 'commercial'
    if s in ('furniture design', 'furniture-design',
             'exhibition design', 'exhibition-design',
             'event design', 'event-design'):
        return 'retail'
    # Fallback: turn into a slug as before
    return s.replace(' ', '-')


def sector_label(slug):
    if not slug:
        return ''
    labels = {
        'residential': 'Residential',
        'retail': 'Retail',
        'commercial': 'Commercial',
        'hospitality': 'Hospitality',
        'urban-planning': 'Urban planning',
        'consultancy': 'Consultancy',
        'institutional': 'Institutional',
    }
    return labels.get(slug, slug.replace('-', ' ').capitalize())


def stage_label(slug):
    labels = {'completed': 'Completed', 'in-progress': 'In progress', 'concept': 'Concept'}
    return labels.get(slug, slug.capitalize() if slug else '')


# ==================== PAGE BUILDERS ====================

def build_home(featured_projects, featured_articles):
    intro_overlay = """
  <!-- ===== SIGNATURE INTRO OVERLAY ===== -->
  <div id="intro-overlay" class="intro-overlay">
    <button class="intro-overlay__skip" aria-label="Skip intro">Skip →</button>
    <svg class="intro-stage" viewBox="0 0 280 220" xmlns="http://www.w3.org/2000/svg">
      <g id="sig-cube-group" stroke="#E8000D" stroke-width="3" fill="none" stroke-linejoin="miter" stroke-miterlimit="10" stroke-linecap="square">
        <circle id="sig-dot" cx="90" cy="90" r="6" fill="#E8000D" stroke="none" style="opacity: 1;"/>
        <path id="sig-line" d="M 90 90 L 190 90" style="opacity: 0;"/>
        <path id="sig-square" d="M 90 90 L 90 190 L 190 190 L 190 90" style="opacity: 0;"/>
        <path id="sig-depth-tl" d="M 90 90 L 130 50" style="opacity: 0;"/>
        <path id="sig-depth-tr" d="M 190 90 L 230 50" style="opacity: 0;"/>
        <path id="sig-depth-br" d="M 190 190 L 230 150" style="opacity: 0;"/>
        <path id="sig-back" d="M 130 50 L 230 50 L 230 150 L 190 190" style="opacity: 0;"/>
      </g>
    </svg>
    <img id="sig-wordmark" src="images/logo.svg" alt="Studio Dotbox" style="opacity: 0; transform: translateY(8px); position: absolute; width: 360px; max-width: 70vw; height: auto;">
  </div>
"""

    # Selected work strip
    work_cards = ""
    if featured_projects:
        for p in featured_projects[:5]:
            img_src = f"images/projects/{p['slug']}/thumb.jpg"
            work_cards += f"""
          <a href="work/{p['slug']}.html" class="home-work__card">
            <div class="home-work__card-image">
              <img src="{img_src}" alt="{normalise_text(p['name'])}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Project image: {p['slug']}/thumb.jpg]';">
            </div>
            <p class="home-work__card-meta"><span>{sector_label(p['sector'])}</span><span>{p['year']}</span></p>
            <p class="home-work__card-title">{normalise_text(p['name'])}</p>
            <p class="home-work__card-stage home-work__card-stage--{p['stage']}">{stage_label(p['stage'])}</p>
          </a>"""
    else:
        # Placeholder cards
        placeholders = [
            ('Residential', 2025, 'completed', 'Completed'),
            ('Retail', 2025, 'in-progress', 'In progress'),
            ('Commercial', 2024, 'completed', 'Completed'),
            ('Hospitality', 2024, 'concept', 'Concept'),
        ]
        for sector, year, stage, stage_text in placeholders:
            work_cards += f"""
          <a href="work.html" class="home-work__card">
            <div class="home-work__card-image">[Project image]</div>
            <p class="home-work__card-meta"><span>{sector}</span><span>{year}</span></p>
            <p class="home-work__card-title">[Project name]</p>
            <p class="home-work__card-stage home-work__card-stage--{stage}">{stage_text}</p>
          </a>"""

    # Journal cards
    journal_cards = ""
    if featured_articles:
        for a in featured_articles[:2]:
            img_src = f"images/journal/{a['slug']}-thumb.jpg"
            journal_cards += f"""
          <a href="journal/{a['slug']}.html" class="home-journal__card">
            <div class="home-journal__card-image">
              <img src="{img_src}" alt="{normalise_text(a['title'])}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Article image: {a['slug']}-thumb.jpg]';">
            </div>
            <p class="home-journal__card-meta"><span>{category_label(a['category'])}</span><span>{a['reading_time']}</span><span>{a['publish_date']}</span></p>
            <p class="home-journal__card-title">{normalise_text(a['title'])}</p>
            <p class="home-journal__card-excerpt">{normalise_text(a['excerpt'])}</p>
          </a>"""
    else:
        journal_cards = """
          <a href="journal.html" class="home-journal__card">
            <div class="home-journal__card-image">[Article image]</div>
            <p class="home-journal__card-meta"><span>Coming soon</span></p>
            <p class="home-journal__card-title">[Next article]</p>
            <p class="home-journal__card-excerpt">More writing from the studio, published regularly.</p>
          </a>"""

    home_body = f"""
  <main>

    <section class="home-hero">
      <div class="home-hero__image">
        <img src="images/hero.jpg" alt="Studio Dotbox" onerror="this.style.display='none'; this.parentElement.innerHTML='[Hero image: 2400 × 1400 px, save as images/hero.jpg]';">
      </div>
      <div class="home-hero__content container">
        <h1 class="home-hero__tagline reveal">Transforming Spaces into <span class="red-accent">Experiences</span>.</h1>
        <p class="home-hero__support reveal">Architecture and interiors, with urban planning at scale. Spaces designed to perform, commercially, functionally, and experientially.</p>
      </div>
    </section>

    <section class="home-intro">
      <div class="container">
        <p class="reveal">Studio Dotbox is a studio-based practice working across homes, retail, workplaces, hospitality, urban planning, and consultancy. Every project is led by the founder, with direct involvement from first conversation to final handover.</p>
      </div>
    </section>

    <section class="home-sectors">
      <div class="container">
        <p class="home-sectors__heading reveal">Where we work</p>
        <ul class="home-sectors__list reveal">
          <li><a href="services.html#residential"><span class="number">01</span>Residential</a></li>
          <li><a href="services.html#retail"><span class="number">02</span>Retail</a></li>
          <li><a href="services.html#commercial"><span class="number">03</span>Commercial</a></li>
          <li><a href="services.html#hospitality"><span class="number">04</span>Hospitality</a></li>
          <li><a href="services.html#urban-planning"><span class="number">05</span>Urban planning</a></li>
          <li><a href="services.html#consultancy"><span class="number">06</span>Consultancy</a></li>
        </ul>
      </div>
    </section>

    <section class="home-work">
      <div class="container">
        <p class="home-work__heading reveal">Selected work</p>
        <div class="home-work__strip reveal">{work_cards}
        </div>
        <div class="home-work__more reveal"><a href="work.html" class="link-arrow">See all work →</a></div>
      </div>
    </section>

    <section class="home-philosophy">
      <div class="container-narrow">
        <p class="home-philosophy__name reveal">Our practice is built on a single idea. We call it Design Out of the Box.</p>
        <p class="home-philosophy__progression reveal">A dot becomes a line. A line becomes a plane. A plane becomes the space you inhabit.</p>
        <a href="philosophy.html" class="link-arrow reveal">Read the philosophy →</a>
      </div>
    </section>

    <section class="home-journal">
      <div class="container">
        <p class="home-journal__heading reveal">From the Journal</p>
        <div class="home-journal__grid reveal">{journal_cards}
        </div>
        <div class="reveal"><a href="journal.html" class="link-arrow">Read more from the studio →</a></div>
      </div>
    </section>

    <section class="home-note">
      <div class="container">
        <p class="home-note__heading reveal">A note from the studio</p>
        <p class="reveal">Studio Dotbox is a small practice. We work closely with a small number of clients each year, on projects we believe in. If that sounds like the kind of conversation you want to have, we would like to hear from you.</p>
      </div>
    </section>

    <section class="home-cta">
      <div class="container-narrow">
        <h2 class="home-cta__heading reveal">Start the conversation.</h2>
        <p class="home-cta__intro reveal">Every project begins with a conversation. Tell us what you are thinking about, whether it is a defined brief or an early idea. No commitment, no pressure, just a chance to understand each other.</p>
        <div class="home-cta__contact reveal">
          <div class="home-cta__contact-item">
            <span class="label">Email</span>
            <a href="mailto:studiodotbox@gmail.com">studiodotbox@gmail.com</a>
          </div>
          <div class="home-cta__contact-item">
            <span class="label">Phone</span>
            <a href="tel:+919571129136">+91 9571 129 136</a>
          </div>
        </div>
        <a href="contact.html" class="btn btn--primary reveal" style="margin-top: var(--space-s);">Send a message →</a>
      </div>
    </section>

  </main>
"""

    write_page("index.html", "Studio Dotbox | Architecture, interiors, and urban planning",
        "Studio Dotbox is a studio-based practice working across homes, retail, workplaces, hospitality, urban planning, and consultancy.",
        home_body, with_animations=True, intro_overlay=intro_overlay,
        extra_stylesheets=["css/home.css"])


def build_studio():
    studio_body = """
  <main>
    <section class="lead">
      <div class="container">
        <h1 class="display reveal">Studio Dotbox is the architectural practice of Pratyush Sharma.</h1>
        <p class="lead__support reveal">A studio-based practice, based in Noida, working across homes, retail, workplaces, hospitality, urban planning, and consultancy.</p>
      </div>
    </section>

    <section class="portrait">
      <div class="container">
        <div class="portrait__image reveal">
          <img src="images/pratyush-portrait.jpg" alt="Pratyush Sharma" onerror="this.style.display='none'; this.parentElement.innerHTML='[Portrait of Pratyush Sharma: 1600 × 900 px, save as images/pratyush-portrait.jpg]';">
        </div>
        <p class="portrait__caption reveal">Pratyush Sharma. Principal Architect and Founder.</p>
      </div>
    </section>

    <section class="background">
      <div class="container">
        <div class="prose reveal">
          <p>I founded Studio Dotbox in 2020, after completing my architectural education and receiving my licence to practice. For the first four years, I worked on the studio alongside full-time roles at established architectural firms. I went full-time on Studio Dotbox in 2024.</p>
          <p>I chose &ldquo;Studio&rdquo; to signal a design-focused practice, not a firm that happens to offer design. &ldquo;Dotbox&rdquo; came from an approach I kept returning to: designing out of the box, for every brief, at every scale.</p>
          <p>In those years, I have designed across a range of typologies: pavilions for Auto Expo, office interiors, private homes, government buildings, hotels, mixed-use projects, retail rollouts executed across India, and luxury residential and commercial work. That variety has shaped how I think about the practice: every project is specific, but the thinking underneath it is consistent.</p>
          <p>Alongside the design work, the studio has built a network of 500+ architects, 50+ contractors, and 150+ vendors across India. For a studio-based practice, that network is as important as the drawings. It is how we move from design to delivery, at any scale.</p>
          <p>I studied at Amity School of Architecture and Planning, Jaipur, from 2014 to 2019. Architecture school changed the way I see the world. I came in curious about buildings and left believing that everything around us is designed: a chair, a spoon, a mobile phone, a street. That shift, from thinking of architecture as a discipline to thinking of it as a way of seeing, is still at the core of how I approach the work.</p>
          <p>Alongside the studio, I completed a Masters in Urban and Regional Planning at VGU, Jaipur, in May 2026. The masters came out of two interests: the first, a curiosity about how cities are designed, which I found myself returning to after architecture; the second, a long-term intention to teach. During the programme, I worked on an urban redevelopment project for several industrial areas in Delhi. Urban planning is now part of how Studio Dotbox thinks about projects at every scale.</p>
        </div>
      </div>
    </section>

    <section class="background" style="border-top: 1px solid var(--color-line);">
      <div class="container">
        <p class="section-label reveal">How I think about the work</p>
        <div class="prose reveal">
          <p>The first move on any project is not a drawing. It is a questionnaire. Before I draw a line, I need to understand what the client actually needs, which is often different from what the brief says. The brief is the client&rsquo;s guess at what they want. The questionnaire is where we get to what they actually need.</p>
          <p>The second move is the site. Orientation, sun, wind, climate, access, topography. These are not poetic considerations. They are the constraints that determine where the building opens, which walls block the heat, how the air moves through the space, and which rooms belong where. By the time I know the site, most of the plan has quietly been decided.</p>
          <p>The third move is where experimentation begins. Form, material, experience. These decisions matter, but they come last in the work, because they are the least constrained. A strong plan in a well-understood site can support many forms. A strong form on a misunderstood site cannot support any.</p>
        </div>
      </div>
    </section>

    <section class="background" style="border-top: 1px solid var(--color-line);">
      <div class="container">
        <p class="section-label reveal">How I work with clients</p>
        <div class="prose reveal">
          <p>I take on a limited number of projects each year, regardless of scale. Every project gets my time in full sittings of three or four hours, not fifteen-minute check-ins between other work. The week is structured around the projects, not the other way around.</p>
          <p>I stay involved with every project from the first conversation to the completed space, and beyond. Once a space is inhabited, the relationship does not end. My relationship with a space I have designed is not temporary, it is permanent.</p>
          <p>I design best with clients who want to be part of the process. Who can tell me how they live, how they move through a day, what a space needs to do for them. The brief is where we start; the conversation is how we get to a design that works.</p>
          <p>I prefer to be involved from the first idea, not brought in to execute a decision that has already been made. Design is the expertise I bring. Execution follows from it.</p>
          <p>Clients can expect regular communication. Weekly during design. Daily during execution. Every decision documented, nothing left to &ldquo;the designer will figure it out.&rdquo;</p>
        </div>
      </div>
    </section>

    <section class="background" style="border-top: 1px solid var(--color-line);">
      <div class="container">
        <p class="section-label reveal">Studio details</p>
        <div class="studio-details reveal">
          <div><p class="studio-details__label">Founded</p><p class="studio-details__value">2020</p></div>
          <div><p class="studio-details__label">Based in</p><p class="studio-details__value">Noida, Delhi NCR</p></div>
          <div><p class="studio-details__label">Email</p><p class="studio-details__value"><a href="mailto:studiodotbox@gmail.com">studiodotbox@gmail.com</a></p></div>
          <div><p class="studio-details__label">Phone</p><p class="studio-details__value"><a href="tel:+919571129136">+91 9571 129 136</a></p></div>
        </div>
      </div>
    </section>

    <section class="continue">
      <div class="container">
        <p class="continue__label reveal">Continue</p>
        <div class="continue__links reveal">
          <a href="work.html" class="link-arrow">See the work →</a>
          <a href="contact.html" class="link-arrow">Start a conversation →</a>
        </div>
      </div>
    </section>
  </main>
"""
    write_page("studio.html", "Studio | Studio Dotbox",
        "Studio Dotbox is the architectural practice of Pratyush Sharma. A studio-based practice based in Noida.",
        studio_body)


def build_philosophy():
    philosophy_body = """
  <main>
    <section id="philosophy-opening" class="philosophy-opening">
      <div class="container-narrow" style="text-align: center;">
        <svg viewBox="0 0 320 260" xmlns="http://www.w3.org/2000/svg" style="display: block; margin: 0 auto;">
          <g id="phil-cube-group" stroke="#E8000D" stroke-width="3" fill="none" stroke-linejoin="miter" stroke-miterlimit="10" stroke-linecap="square">
            <circle id="phil-dot" cx="100" cy="110" r="8" fill="#E8000D" stroke="none" style="opacity: 0;"/>
            <path id="phil-line" d="M 100 110 L 220 110" style="opacity: 0;"/>
            <path id="phil-square" d="M 100 110 L 100 230 L 220 230 L 220 110" style="opacity: 0;"/>
            <path id="phil-depth-tl" d="M 100 110 L 160 50" style="opacity: 0;"/>
            <path id="phil-depth-tr" d="M 220 110 L 280 50" style="opacity: 0;"/>
            <path id="phil-depth-br" d="M 220 230 L 280 170" style="opacity: 0;"/>
            <path id="phil-back" d="M 160 50 L 280 50 L 280 170 L 220 230" style="opacity: 0;"/>
          </g>
        </svg>
        <p id="philosophy-text" class="philosophy-stages" style="opacity: 0; transform: translateY(20px); margin-top: var(--space-l);">
          <span>Dot.</span> Line. Plane. Box.
        </p>
      </div>
    </section>

    <section style="padding: var(--space-m) 0;">
      <div class="container-narrow">
        <p class="thesis reveal">We call this Design Out of the Box.</p>
        <p class="reveal" style="text-align: center; font-size: var(--size-body-large); color: var(--color-text-muted); max-width: 640px; margin: 0 auto;">It is how we think about every project, from the first conversation to the final handover. A single idea, extended through four stages, into the space you inhabit.</p>
      </div>
    </section>

    <section class="stages">
      <div class="container">
        <div class="stage reveal">
          <div class="stage__mark"><svg viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg"><circle class="fade-in" cx="40" cy="40" r="6" fill="#E8000D"/></svg></div>
          <div class="stage__content">
            <h3>The Dot.</h3>
            <p class="stage__sub">The starting point.</p>
            <p>The dot is the brief. Before any drawing begins, we establish what the space needs to do, who it is for, and what will define its success. Budget, site, function, client, context. Every project begins here, with a clear, agreed starting point.</p>
          </div>
        </div>
        <div class="stage reveal">
          <div class="stage__mark"><svg viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg"><line class="draw-stroke" x1="10" y1="40" x2="70" y2="40" stroke="#E8000D" stroke-width="3" stroke-linecap="square"/></svg></div>
          <div class="stage__content">
            <h3>The Line.</h3>
            <p class="stage__sub">The direction.</p>
            <p>The line is the plan. The circulation, the adjacencies, the flow from one space to the next. At this stage we resolve how the space works before we consider how it looks. A good plan makes the rest of the work easier. A poor plan cannot be saved by materials.</p>
          </div>
        </div>
        <div class="stage reveal">
          <div class="stage__mark"><svg viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg"><rect class="draw-stroke" x="15" y="15" width="50" height="50" fill="none" stroke="#E8000D" stroke-width="3"/></svg></div>
          <div class="stage__content">
            <h3>The Plane.</h3>
            <p class="stage__sub">The surface.</p>
            <p>The plane is the material decision. Floors, walls, ceilings, finishes, light, texture. This is where the design becomes tangible, and where most of the budget is spent. Every choice at this stage is judged against two questions: does it serve the space, and will it hold up over time.</p>
          </div>
        </div>
        <div class="stage reveal">
          <div class="stage__mark">
            <svg viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
              <g fill="none" stroke="#E8000D" stroke-width="3" stroke-linejoin="miter" stroke-linecap="square">
                <rect class="fade-in" x="15" y="25" width="40" height="40"/>
                <line class="fade-in-delayed-1" x1="15" y1="25" x2="30" y2="10"/>
                <line class="fade-in-delayed-1" x1="55" y1="25" x2="70" y2="10"/>
                <line class="fade-in-delayed-1" x1="55" y1="65" x2="70" y2="50"/>
                <path class="fade-in-delayed-2" d="M 30 10 L 70 10 L 70 50"/>
                <line class="fade-in-delayed-3" x1="70" y1="50" x2="55" y2="65"/>
              </g>
            </svg>
          </div>
          <div class="stage__content">
            <h3>The Box.</h3>
            <p class="stage__sub">The inhabited space.</p>
            <p>The box is the finished project. But the work is not finished when construction ends. It is finished when the space performs. Commercially for retail and hospitality. Functionally for homes and workplaces. Experientially across all of them. The box is judged by use, not by photographs.</p>
          </div>
        </div>
      </div>
    </section>

    <section class="practice">
      <div class="container">
        <p class="practice__heading reveal">What this means in practice</p>
        <div class="practice__lines reveal">
          <p>We begin every project by locating the dot. The single idea the space needs to hold.</p>
          <p>We design for how the space performs, not just how it photographs.</p>
          <p>We treat material, light, and proportion as the plane: the physical ground the experience stands on.</p>
          <p>We believe the final box is judged by the people who inhabit it, not by the people who design it.</p>
        </div>
      </div>
    </section>

    <section class="section">
      <div class="container-narrow">
        <p class="philosophy-closing reveal">Most design starts with a box. We start with a dot, and we work outward.</p>
      </div>
    </section>

    <section class="continue">
      <div class="container">
        <p class="continue__label reveal">Continue</p>
        <div class="continue__links reveal">
          <a href="how-we-work.html" class="link-arrow">See how we work →</a>
          <a href="work.html" class="link-arrow">See the projects →</a>
        </div>
      </div>
    </section>
  </main>
"""
    write_page("philosophy.html", "Philosophy | Studio Dotbox",
        "Design Out of the Box is how Studio Dotbox thinks about every project.",
        philosophy_body, with_animations=True)


def build_services():
    def sector_block(num, anchor, name, description, items, link_text):
        items_html = "\n".join(f"          <li>{item}</li>" for item in items)
        return f"""
    <section id="{anchor}" class="sector reveal">
      <div class="container">
        <p class="sector__number">{num}</p>
        <h2 class="sector__title">{name}</h2>
        <p class="sector__description">{description}</p>
        <p class="sector__included">What&rsquo;s typically included</p>
        <ul class="sector__list">
{items_html}
        </ul>
        <a href="work.html" class="link-arrow">{link_text} →</a>
      </div>
    </section>
"""

    body = """
  <main>
    <section class="lead">
      <div class="container">
        <h1 class="display reveal">What we do, and who we do it for.</h1>
        <p class="lead__support reveal">Studio Dotbox works across six disciplines. The approach is consistent. The context is different every time.</p>
      </div>
    </section>
"""
    body += sector_block("01", "residential", "Residential",
        "Homes are not projects, they are places. We design homes that hold up across years of use, weather, furniture, and lives. Every decision is made with the people who will live in the space in mind, because the space will belong to them long after we hand over the keys. The test of a residential project is not the handover photograph, it is the tenth year.",
        ["Architectural design", "Interior design and decor", "Furniture design", "Services planning", "Material selection", "Vendor and contractor coordination", "Site supervision", "Handover and post-completion support"],
        "See residential projects")
    body += sector_block("02", "retail", "Retail",
        "Retail design is a commercial decision first, and a design decision second. We design retail spaces that perform across brand expression, footfall behaviour, and conversion, whether the brief is a single flagship or a rollout across multiple cities. The purpose of a retail space is not to look right, it is to work. We design for both.",
        ["Concept and design development", "Brand integration", "Store layout and customer journey", "Material and fixture specification", "Rollout design and standardisation", "On-site coordination and execution"],
        "See retail projects")
    body += sector_block("03", "commercial", "Commercial",
        "A workplace is judged by the work that gets done inside it. We design office and commercial interiors around how the people inside them actually spend their days. Meeting rooms that get used. Corridors that connect. Quiet spaces that stay quiet. The brand on the wall matters less than the behaviour in the room.",
        ["Space planning and circulation", "Interior design", "Meeting, collaboration, interaction, and quiet zones", "Lighting, acoustics, and building services", "Material selection", "Branding integration", "Site coordination"],
        "See commercial projects")
    body += sector_block("04", "hospitality", "Hospitality",
        "A hospitality space is remembered or it is forgotten. There is no middle ground. We design restaurants, cafes, hotels, and guest spaces around how a guest moves through them, what they feel at each stage, and what brings them back. The interior is part of it. The journey through it matters more.",
        ["Concept and design development", "Guest journey design", "Interior design", "Material selection", "Services planning", "Back-of-house planning", "Execution coordination"],
        "See hospitality projects")
    body += sector_block("05", "urban-planning", "Urban planning",
        "The scale of the city is a different kind of design problem, but the thinking is the same. We bring urban planning to masterplans, mixed-use developments, campus design, and urban redevelopment, integrating architecture, interiors, and planning as one continuous body of work. A good home belongs to a good street. A good office belongs to a neighbourhood. We design at every scale because the scales are connected.",
        ["Master planning", "Township development", "Mixed-use and campus planning", "Urban redevelopment strategy and planning", "Stakeholder surveys", "Phased implementation planning", "Integration with architecture and interiors"],
        "See urban planning projects")
    body += sector_block("06", "consultancy", "Consultancy",
        "Sometimes the work is not designing the space. It is helping a client decide what to build, validating a brief before it commits to a budget, auditing a building for accessibility or structure, finding the right team to deliver the project, or steering execution through to completion. Studio Dotbox offers its expertise in all of these forms, without requiring a full design engagement. The studio&rsquo;s value is in the thinking, not only the drawings.",
        ["Strategic advisory", "Physical accessibility audits and design", "Structural audits", "Feasibility assessments", "Project brief development and validation", "Site and space evaluation", "Concept reviews", "Material and vendor strategy", "Talent acquisition (architects, designers, and project teams)", "Project management consultancy"],
        "See consultancy projects")

    body += """
    <section class="cta">
      <div class="container-narrow">
        <p class="cta__intro reveal">If this sounds like the kind of practice you want to work with, we would like to hear from you.</p>
        <a href="contact.html" class="btn btn--primary reveal">Start the conversation →</a>
      </div>
    </section>
  </main>
"""
    write_page("services.html", "Services | Studio Dotbox",
        "Studio Dotbox works across six disciplines: residential, retail, commercial, hospitality, urban planning, and consultancy.",
        body)


def build_how_we_work():
    def stage_detail(num, title, paragraphs, receive, need):
        paras_html = "\n        ".join(f"<p>{p}</p>" for p in paragraphs)
        receive_html = "\n            ".join(f"<li>{i}</li>" for i in receive)
        need_html = "\n            ".join(f"<li>{i}</li>" for i in need)
        return f"""
    <section class="stage-detail reveal">
      <div class="container">
        <h2 class="stage-detail__title">{num}. {title}</h2>
        {paras_html}
        <div class="stage-detail__lists">
          <div><h4>What you receive</h4><ul>{receive_html}</ul></div>
          <div><h4>What we need from you</h4><ul>{need_html}</ul></div>
        </div>
      </div>
    </section>
"""

    body = """
  <main>
    <section class="lead">
      <div class="container">
        <h1 class="display reveal">How we work.</h1>
        <p class="lead__support reveal">From the first conversation to handing over the keys, this is what working with Studio Dotbox looks like.</p>
      </div>
    </section>

    <section class="process-sequence">
      <div class="container">
        <div class="process-step reveal"><div class="process-step__number">01</div><div><p class="process-step__name">The conversation.</p><p class="process-step__description">Understanding the brief, the site, the budget, and the client.</p></div></div>
        <div class="process-step reveal"><div class="process-step__number">02</div><div><p class="process-step__name">The proposal.</p><p class="process-step__description">A detailed proposal covering scope, fees, stages, and deliverables.</p></div></div>
        <div class="process-step reveal"><div class="process-step__number">03</div><div><p class="process-step__name">The design.</p><p class="process-step__description">Concept, development, and detailed design, in clear stages.</p></div></div>
        <div class="process-step reveal"><div class="process-step__number">04</div><div><p class="process-step__name">The build.</p><p class="process-step__description">Site supervision, coordination, and execution.</p></div></div>
        <div class="process-step reveal"><div class="process-step__number">05</div><div><p class="process-step__name">The handover.</p><p class="process-step__description">Final walkthrough, documentation, and the space in use.</p></div></div>
      </div>
    </section>
"""
    body += stage_detail("01", "The conversation.",
        ["Every project begins with a conversation. It is how we understand whether the project is right for the studio, and whether the studio is right for the project.",
         "At this stage, we discuss what you are thinking about, the site or space in question, the brief as you see it, the timeline, and the budget. It is a two-way conversation, not a sales call."],
        ["A clear understanding of how we work", "An honest view of whether Studio Dotbox is the right fit", "A proposed next step if we both want to go forward"],
        ["A willingness to share the context of the project", "Availability for one or two initial conversations", "Honesty about timeline, budget, and expectations"])
    body += stage_detail("02", "The proposal.",
        ["Once we agree the project is a fit, we share a detailed proposal. This is where the commercial relationship begins, and where we establish a shared understanding of what the work will be.",
         "The proposal covers scope, fees, stages, deliverables, and timeline. Every term is written plainly, with no ambiguity. You sign nothing until you have read and understood the proposal in full."],
        ["A written proposal document", "A clear fee structure with staged payments", "A list of deliverables for each stage", "A proposed project timeline"],
        ["Review and approval of the proposal", "A signed agreement to begin work", "The initial payment that triggers the design stage"])
    body += stage_detail("03", "The design.",
        ["The design stage is where the space is resolved. It is the longest stage and the most collaborative. We work through concept, development, and detail, in clear phases, with sign-off at each one.",
         "This stage is where most of the thinking happens. Site studies, concept explorations, planning resolution, material choices, services planning, and detailed drawings all sit here."],
        ["Concept drawings and mood boards", "Developed design drawings", "Detailed design and technical drawings", "Material and finish proposals", "Regular updates and review meetings"],
        ["Availability for review meetings", "Clear feedback on concept, design, and material decisions", "Sign-off at the end of each design phase"])
    body += stage_detail("04", "The build.",
        ["Once the design is approved, the project moves to site. At this stage, we work closely with contractors, vendors, and site teams to ensure the space is built the way it was designed.",
         "Site supervision, contractor coordination, vendor management, and day-to-day problem-solving all happen here. We stay present on site, because buildings are made during the build, not during the design."],
        ["Daily updates during active site work", "Weekly progress meetings", "Resolution of site-level design decisions", "Coordination of all contractors and vendors"],
        ["Availability for site meetings as needed", "Timely decisions when site conditions require them", "Staged payments released as milestones are reached"])
    body += stage_detail("05", "The handover.",
        ["The handover is not the end of the project. It is the beginning of the space being used. We walk through the completed space with you, hand over all documentation, and resolve any final items. After that, we stay available as the space is lived in."],
        ["A final walkthrough of the completed space", "All drawings, documents, and warranties", "A snag list resolution process", "Ongoing availability for questions once the space is in use"],
        ["Time for the walkthrough", "Honest feedback on the finished space", "Final payment on handover"])

    body += """
    <section class="cta">
      <div class="container-narrow">
        <p class="cta__intro reveal">If you are thinking about a project, the first step is a conversation.</p>
        <a href="contact.html" class="btn btn--primary reveal">Start the conversation →</a>
      </div>
    </section>
  </main>
"""
    write_page("how-we-work.html", "How we work | Studio Dotbox",
        "From the first conversation to handing over the keys, this is what working with Studio Dotbox looks like.",
        body)


def build_contact():
    body = """
  <main>
    <section class="lead">
      <div class="container">
        <h1 class="display reveal">Start the conversation.</h1>
        <p class="lead__support reveal">Every project begins with a conversation. Tell us what you are thinking about, whether it is a defined brief or an early idea. There is no commitment, no pressure, just a chance to understand each other.</p>
      </div>
    </section>

    <section style="padding: var(--space-m) 0;">
      <div class="container" style="max-width: 720px;">
        <form id="contact-form" class="reveal" action="https://api.web3forms.com/submit" method="POST">
          <input type="hidden" name="access_key" value="YOUR_WEB3FORMS_KEY_HERE">
          <input type="hidden" name="subject" value="New enquiry from Studio Dotbox website">
          <input type="checkbox" name="botcheck" style="display:none;">

          <div class="form-field"><label for="name">Your name <span class="required">*</span></label><input type="text" id="name" name="name" placeholder="Full name" required></div>
          <div class="form-field"><label for="email">Email <span class="required">*</span></label><input type="email" id="email" name="email" placeholder="xyz@example.com" required></div>
          <div class="form-field"><label for="phone">Phone</label><input type="tel" id="phone" name="phone" placeholder="+00 0000 000 000"></div>
          <div class="form-field"><label for="project-type">What kind of project are you thinking about? <span class="required">*</span></label>
            <select id="project-type" name="project-type" required>
              <option value="">Select an option</option>
              <option>Residential</option><option>Retail</option><option>Commercial</option>
              <option>Hospitality</option><option>Urban planning</option><option>Consultancy</option>
              <option>Other</option><option>Not sure yet</option>
            </select>
          </div>
          <div class="form-field"><label for="location">Location of the project</label><input type="text" id="location" name="location" placeholder="Area, City, State, Country"></div>
          <div class="form-field"><label for="timeline">When are you hoping to start?</label>
            <select id="timeline" name="timeline">
              <option value="">Select an option</option>
              <option>Within 3 months</option><option>3 to 6 months</option><option>6 to 12 months</option>
              <option>Later than that</option><option>Still exploring</option>
            </select>
          </div>
          <div class="form-field"><label for="message">Tell us a little about what you have in mind <span class="required">*</span></label>
            <textarea id="message" name="message" placeholder="Just tell us whatever there is in your mind" required></textarea>
            <p class="help-text">A few sentences is plenty. We will ask more when we talk.</p>
          </div>
          <div class="form-field"><label for="referral">How did you find us?</label>
            <select id="referral" name="referral">
              <option value="">Select an option</option>
              <option>Google Search</option><option>LinkedIn</option><option>Instagram</option>
              <option>YouTube</option><option>Pinterest</option><option>Referral</option><option>Other</option>
            </select>
          </div>
          <button type="submit" class="btn btn--primary">Start the conversation →</button>
        </form>

        <div id="contact-success" style="display: none; padding: var(--space-l) 0; text-align: center;">
          <h2 class="h1" style="margin-bottom: var(--space-s);">Note received.</h2>
          <p class="body-large" style="margin-bottom: var(--space-m); color: var(--color-text-muted);">We will be in touch within two working days.</p>
          <a href="index.html" class="link-arrow">Back to the studio →</a>
        </div>
      </div>
    </section>

    <section style="padding: var(--space-m) 0; border-top: 1px solid var(--color-line);">
      <div class="container" style="max-width: 720px;">
        <p class="section-label reveal">After you send this</p>
        <p class="reveal" style="font-size: var(--size-body-large); font-weight: var(--weight-light); line-height: 1.6; max-width: 720px;">We get back to you within two working days. The first step is usually a short call, where we ask a few questions and you can ask a few of yours. If we are both happy to go further, we visit the site (or arrange a meeting with you) and share a proposal within a week of that.</p>
      </div>
    </section>

    <section style="padding: var(--space-m) 0; border-top: 1px solid var(--color-line);">
      <div class="container" style="max-width: 720px;">
        <p class="section-label reveal">Or reach us directly</p>
        <div class="reveal" style="display: grid; gap: 12px; font-size: var(--size-body); font-weight: var(--weight-light);">
          <p>Email: <a href="mailto:studiodotbox@gmail.com" class="link-arrow" style="display: inline;">studiodotbox@gmail.com</a></p>
          <p>Phone: <a href="tel:+919571129136" class="link-arrow" style="display: inline;">+91 9571 129 136</a></p>
          <p style="color: var(--color-text-muted); margin-top: 12px;">Based in Noida. Meetings on request.</p>
        </div>
      </div>
    </section>

    <section class="quiet-note reveal">Thank you for considering Studio Dotbox.</section>
  </main>

  <script>
    (function() {
      var form = document.getElementById('contact-form');
      var success = document.getElementById('contact-success');
      if (!form) return;
      form.addEventListener('submit', function(e) {
        e.preventDefault();
        var data = new FormData(form);
        var ak = data.get('access_key') || '';
        if (!ak || ak.indexOf('YOUR_') === 0) {
          alert('Form preview only. Add your Web3Forms access key to enable submission.');
          return;
        }
        fetch('https://api.web3forms.com/submit', { method: 'POST', body: data })
          .then(function(r) {
            if (r.ok) {
              form.style.display = 'none';
              success.style.display = 'block';
              window.scrollTo({ top: form.offsetTop - 100, behavior: 'smooth' });
            } else { alert('Sorry, something went wrong. Please try email instead.'); }
          })
          .catch(function() { alert('Sorry, something went wrong. Please try email instead.'); });
      });
    })();
  </script>
"""
    write_page("contact.html", "Contact | Studio Dotbox",
        "Start the conversation with Studio Dotbox. Every project begins with a conversation.",
        body)


def build_work_with_us():
    body = """
  <main>
    <section class="lead">
      <div class="container">
        <h1 class="display reveal">Work with us.</h1>
        <p class="lead__support reveal">Studio Dotbox grows through the network it builds. Professionals, contractors, and vendors who work with us well, once, often become part of how we deliver the next project. If you are looking to join that network, this is where you start.</p>
      </div>
    </section>

    <section id="professionals" class="sector reveal">
      <div class="container">
        <p class="sector__number">For professionals</p>
        <h2 class="sector__title">For architects, designers, and freelancers.</h2>
        <p class="sector__description">Studio Dotbox collaborates with a growing network of design professionals across India, on projects spanning residential, retail, commercial, hospitality, urban planning, and consultancy. Architects, designers, BIM specialists, freelancers, and other collaborators, if you are looking for project collaborations, we would like to know you.</p>
        <p class="sector__description">The registration is a short form. It takes about five minutes. We review every submission personally, and we get back to you within two working days if there is a current fit.</p>
        <a href="https://forms.gle/w5UwMM72afXVAWij7" target="_blank" rel="noopener" class="btn btn--primary">Register as a professional →</a>
      </div>
    </section>

    <section id="contractors" class="sector reveal">
      <div class="container">
        <p class="sector__number">For contractors</p>
        <h2 class="sector__title">For execution partners and turnkey teams.</h2>
        <p class="sector__description">Studio Dotbox works with a trusted network of contractors, execution partners, turnkey teams, and site teams across India. On projects of every scale, from single-site residential builds to retail rollouts across multiple cities. If you have a team, a track record, and the capacity to take on work with a studio that documents every decision, we would like to know you.</p>
        <a href="https://forms.gle/67Vmwcd8GswKKGTd6" target="_blank" rel="noopener" class="btn btn--primary">Apply for empanelment →</a>
      </div>
    </section>

    <section id="vendors" class="sector reveal">
      <div class="container">
        <p class="sector__number">For vendors</p>
        <h2 class="sector__title">For suppliers, manufacturers, and product companies.</h2>
        <p class="sector__description">Studio Dotbox works with a network of vendors, suppliers, manufacturers, brands, and product companies across materials, finishes, lighting, furniture, fittings, and everything else that makes a designed space possible. If you want your products to be specified on projects across six disciplines, we would like to know your catalogue.</p>
        <a href="https://forms.gle/ux1daK99iSQHy9MG6" target="_blank" rel="noopener" class="btn btn--primary">Apply for empanelment →</a>
      </div>
    </section>

    <section class="sector reveal">
      <div class="container">
        <p class="section-label">A note on how we work</p>
        <div class="prose">
          <p>We treat the network as part of the studio. We pay on time. We credit collaborators where credit is due. We do not ask for unpaid work. And we stay in touch, whether or not there is a live project, because that is how a studio-based practice grows.</p>
        </div>
      </div>
    </section>
  </main>
"""
    write_page("work-with-us.html", "Work with us | Studio Dotbox",
        "Studio Dotbox grows through the network it builds.", body)


def _legal_body(heading, updated, sections):
    """Shared layout for Terms and Privacy. sections = list of (title, paragraph)."""
    blocks = ""
    for title, para in sections:
        blocks += f"""
      <div class="reveal" style="margin-bottom: var(--space-l);">
        <h2 class="h3" style="margin-bottom: var(--space-s);">{title}</h2>
        <p class="body-large" style="color: var(--color-text-muted); font-weight: var(--weight-light); line-height: 1.7;">{para}</p>
      </div>"""
    return f"""
  <main>
    <section style="padding: var(--space-xl) 0 var(--space-l);">
      <div class="container-narrow">
        <p class="section-label reveal">{updated}</p>
        <h1 class="h1 reveal" style="margin-bottom: var(--space-l);">{heading}</h1>
        {blocks}
        <div class="reveal" style="margin-top: var(--space-l);">
          <a href="index.html" class="link-arrow">Back to the studio →</a>
        </div>
      </div>
    </section>
  </main>
"""


def build_terms():
    sections = [
        ("Using this site", "This website presents the work, writing, and services of Studio Dotbox. You are welcome to read, reference, and share it. By using the site you agree to use it lawfully and not to disrupt, copy at scale, or misrepresent its contents."),
        ("Our work and content", "All text, drawings, photographs, and design on this site belong to Studio Dotbox or are used with permission, unless stated otherwise. Please do not reproduce them for commercial use without asking us first. We are usually happy to say yes."),
        ("Enquiries and no obligation", "Sending an enquiry through this site does not create a contract or a professional engagement. Any project relationship begins only through a separate written agreement signed by both parties."),
        ("Accuracy", "We keep this site as accurate as we reasonably can, but content may change without notice. We do not warrant that every detail is current or error free, and we are not liable for decisions made solely on the basis of information here."),
        ("Governing law", "These terms are governed by the laws of India, with jurisdiction in the courts of Uttar Pradesh."),
        ("Contact", "Questions about these terms can go to studiodotbox@gmail.com."),
    ]
    body = _legal_body("Terms of use", "Last updated July 2026", sections)
    write_page("terms.html", "Terms of use | Studio Dotbox",
        "The terms that govern use of the Studio Dotbox website.", body, with_animations=True)


def build_privacy():
    sections = [
        ("What we collect", "If you contact us through the enquiry or subscribe forms, we receive the details you choose to share: your name, email, phone number, and anything you write about your project. We do not run trackers that build a profile of you beyond standard, aggregate visit data."),
        ("How we use it", "We use your details only to respond to you, understand your enquiry, and, if you subscribe, to send occasional writing from the studio. We do not sell or rent your information to anyone."),
        ("Form handling", "Our forms are processed through Web3Forms, a third-party service that passes your submission to our inbox. Your data is handled according to their processing terms in transit; we hold it only as long as needed to reply and keep a record of the conversation."),
        ("Your choices", "You can ask us to show, correct, or delete the information we hold about you, or to stop sending you updates, at any time. Just email us and we will act on it promptly."),
        ("Contact", "For anything about your privacy, write to studiodotbox@gmail.com."),
    ]
    body = _legal_body("Privacy", "Last updated July 2026", sections)
    write_page("privacy.html", "Privacy | Studio Dotbox",
        "How Studio Dotbox handles the information you share with us.", body, with_animations=True)


def build_404():
    body = """
  <main>
    <section style="min-height: 60vh; display: flex; align-items: center; justify-content: center; padding: var(--space-l) 0;">
      <div class="container-narrow" style="text-align: center;">
        <div class="reveal" style="margin-bottom: var(--space-m);">
          <svg id="failed-mark" width="120" height="60" viewBox="0 0 120 60" xmlns="http://www.w3.org/2000/svg" style="display: block; margin: 0 auto;">
            <circle cx="60" cy="30" r="6" fill="#E8000D"/>
            <line id="failed-line" x1="60" y1="30" x2="60" y2="30" stroke="#E8000D" stroke-width="3" stroke-linecap="round"/>
          </svg>
        </div>
        <h1 class="h1 reveal" style="margin-bottom: var(--space-s);">This page does not exist.</h1>
        <p class="body-large reveal" style="color: var(--color-text-muted); margin-bottom: var(--space-l); max-width: 540px; margin-left: auto; margin-right: auto;">Something did not resolve. Let us get you back to somewhere that does.</p>
        <div class="continue__links reveal" style="justify-content: center;">
          <a href="index.html" class="link-arrow">Home →</a>
          <a href="work.html" class="link-arrow">See the work →</a>
          <a href="contact.html" class="link-arrow">Start the conversation →</a>
        </div>
      </div>
    </section>
  </main>
"""
    write_page("404.html", "Page not found | Studio Dotbox",
        "This page does not exist on Studio Dotbox.", body, with_animations=True)


def build_work_archive(projects):
    """Build the Work archive page from the project list."""
    if not projects:
        # Use a default placeholder grid if no projects in Excel
        cards = """
          <a href="work.html" class="work-card" data-sector="residential" data-stage="completed">
            <div class="work-card__image">[Project image]</div>
            <p class="work-card__meta"><span>Residential</span><span>2025</span></p>
            <p class="work-card__title">[Project name]</p>
            <p class="work-card__stage">Completed</p>
          </a>
"""
    else:
        cards = ""
        for p in projects:
            img_src = f"images/projects/{p['slug']}/thumb.jpg"
            cards += f"""
          <a href="work/{p['slug']}.html" class="work-card" data-sector="{p['sector']}" data-stage="{p['stage']}">
            <div class="work-card__image">
              <img src="{img_src}" alt="{normalise_text(p['name'])}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Project image]';">
            </div>
            <p class="work-card__meta"><span>{sector_label(p['sector'])}</span><span>{p['year']}</span></p>
            <p class="work-card__title">{normalise_text(p['name'])}</p>
            <p class="work-card__stage work-card__stage--{p['stage']}">{stage_label(p['stage'])}</p>
          </a>
"""

    body = f"""
  <main>
    <section class="lead">
      <div class="container">
        <h1 class="display reveal">A record of the studio&rsquo;s work.</h1>
        <p class="lead__support reveal">Completed projects, projects in progress, and work on the board. Every project is an answer to a specific question. These are ours.</p>
      </div>
    </section>

    <section style="padding: var(--space-s) 0;">
      <div class="container">
        <div class="filters reveal">
          <div class="filters__group">
            <span class="filters__label">By sector:</span>
            <button class="filters__btn active" data-group="sector" data-value="all">All</button>
            <button class="filters__btn" data-group="sector" data-value="residential">Residential</button>
            <button class="filters__btn" data-group="sector" data-value="retail">Retail</button>
            <button class="filters__btn" data-group="sector" data-value="commercial">Commercial</button>
            <button class="filters__btn" data-group="sector" data-value="hospitality">Hospitality</button>
            <button class="filters__btn" data-group="sector" data-value="urban-planning">Urban planning</button>
            <button class="filters__btn" data-group="sector" data-value="consultancy">Consultancy</button>
            <button class="filters__btn" data-group="sector" data-value="institutional">Institutional</button>
          </div>
        </div>
        <div class="filters reveal">
          <div class="filters__group">
            <span class="filters__label">By stage:</span>
            <button class="filters__btn active" data-group="stage" data-value="all">All</button>
            <button class="filters__btn" data-group="stage" data-value="completed">Completed</button>
            <button class="filters__btn" data-group="stage" data-value="in-progress">In progress</button>
            <button class="filters__btn" data-group="stage" data-value="concept">Concept</button>
          </div>
        </div>
      </div>
    </section>

    <section>
      <div class="container">
        <div class="work-grid reveal">{cards}
        </div>
      </div>
    </section>

    <section class="cta">
      <div class="container-narrow">
        <p class="cta__intro reveal">If you are considering a project, we would like to hear about it.</p>
        <a href="contact.html" class="btn btn--primary reveal">Start a conversation →</a>
      </div>
    </section>
  </main>
"""
    write_page("work.html", "Work | Studio Dotbox",
        "A record of Studio Dotbox's work.", body)


def build_project_page(project):
    """Build an individual project page."""
    p = project
    narrative_html = paragraphs_to_html(p['narrative'])

    # Gallery
    gallery_html = ""
    n = p['gallery_count']
    if n > 0:
        i = 1
        # Pair images two-by-two if even, otherwise first image full-width
        if n % 2 == 1:
            # Single full-width first image
            gallery_html += f"""
        <div class="project-gallery__image">
          <img src="../images/projects/{p['slug']}/{i:02d}.jpg" alt="{normalise_text(p['name'])} {i}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Image {i:02d}.jpg]';">
        </div>"""
            i += 1
        # Remaining images in pairs
        while i < n:
            gallery_html += f"""
        <div class="project-gallery__pair">
          <div class="project-gallery__image">
            <img src="../images/projects/{p['slug']}/{i:02d}.jpg" alt="{normalise_text(p['name'])} {i}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Image {i:02d}.jpg]';">
          </div>
          <div class="project-gallery__image">
            <img src="../images/projects/{p['slug']}/{i+1:02d}.jpg" alt="{normalise_text(p['name'])} {i+1}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Image {(i+1):02d}.jpg]';">
          </div>
        </div>"""
            i += 2
        if i == n:
            gallery_html += f"""
        <div class="project-gallery__image">
          <img src="../images/projects/{p['slug']}/{i:02d}.jpg" alt="{normalise_text(p['name'])} {i}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Image {i:02d}.jpg]';">
        </div>"""

    # Project details (only show if there's at least one filled in)
    details_html = ""
    has_details = p['materials'] or p['floor_area'] or p['period'] or p['collaborators']
    if has_details:
        details_pairs = []
        if p['materials']:
            details_pairs.append(('Materials', p['materials']))
        if p['floor_area']:
            details_pairs.append(('Floor area', p['floor_area']))
        if p['period']:
            details_pairs.append(('Construction period', p['period']))
        if p['collaborators']:
            details_pairs.append(('Collaborators', p['collaborators']))

        details_inner = ""
        for label, value in details_pairs:
            details_inner += f"""
          <div>
            <p class="project-details-list__label">{label}</p>
            <p>{normalise_text(value)}</p>
          </div>"""

        details_html = f"""
    <section style="padding: var(--space-m) 0; border-top: 1px solid var(--color-line);">
      <div class="container-narrow">
        <p class="section-label reveal">Project details</p>
        <div class="project-details-list reveal">{details_inner}
        </div>
      </div>
    </section>
"""

    body = f"""
  <main>
    <section style="padding: 0;">
      <div class="container" style="padding-top: var(--space-s);">
        <div class="project-hero reveal">
          <img src="../images/projects/{p['slug']}/hero.jpg" alt="{normalise_text(p['name'])}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Hero image: 2400 × 1350 px, save as images/projects/{p['slug']}/hero.jpg]';">
        </div>
      </div>
    </section>

    <section class="project-meta">
      <div class="container">
        <h1 class="project-meta__title reveal">{normalise_text(p['name'])}</h1>
        <p class="project-meta__description reveal">{normalise_text(p['description'])}</p>
        <div class="project-meta__details reveal">
          <span>Sector: {sector_label(p['sector'])}</span>
          <span>Stage: {stage_label(p['stage'])}</span>
          <span>Year: {p['year']}</span>
          <span>Location: {normalise_text(p['location'])}</span>
        </div>
      </div>
    </section>

    <section style="padding: var(--space-m) 0;">
      <div class="container-narrow">
        <div class="prose reveal">
{narrative_html}
        </div>
      </div>
    </section>

    <section style="padding: 0 0 var(--space-m);">
      <div class="container">
        <div class="project-gallery reveal">{gallery_html}
        </div>
      </div>
    </section>
{details_html}
    <section class="continue">
      <div class="container">
        <p class="continue__label reveal">Next</p>
        <div class="continue__links reveal">
          <a href="../work.html" class="link-arrow">All projects →</a>
          <a href="../contact.html" class="link-arrow">Start a conversation →</a>
        </div>
      </div>
    </section>
  </main>
"""
    write_page(f"work/{p['slug']}.html",
        f"{p['name']} | Studio Dotbox",
        normalise_text(p['description']) or f"A Studio Dotbox project: {p['name']}.",
        body, depth=1)


def build_journal_archive(articles):
    if not articles:
        body = """
  <main>
    <section class="lead">
      <div class="container">
        <h1 class="display reveal">The Journal.</h1>
        <p class="lead__support reveal">Notes on architecture, design, and the practice. Written by Studio Dotbox.</p>
      </div>
    </section>
    <section style="padding: var(--space-m) 0; border-top: 1px solid var(--color-line);">
      <div class="container">
        <p class="section-label reveal">More writing, on its way.</p>
        <p class="reveal" style="font-size: var(--size-body-large); font-weight: var(--weight-light); color: var(--color-text-muted); max-width: 640px;">New articles are published regularly. Subscribe in the footer below to receive new writing as it comes out.</p>
      </div>
    </section>
  </main>
"""
    else:
        # Featured article (first in list, or article marked featured)
        featured = next((a for a in articles if a['featured']), articles[0])
        # All articles excluding featured
        recent = [a for a in articles if a['slug'] != featured['slug']]

        recent_html = ""
        for a in recent:
            recent_html += f"""
          <a href="journal/{a['slug']}.html" class="journal-card">
            <div class="journal-card__image">
              <img src="images/journal/{a['slug']}-thumb.jpg" alt="{normalise_text(a['title'])}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Article image]';">
            </div>
            <p class="journal-card__meta"><span>{category_label(a['category'])}</span><span>{a['reading_time']}</span><span>{a['publish_date']}</span></p>
            <p class="journal-card__title">{normalise_text(a['title'])}</p>
            <p class="journal-card__excerpt">{normalise_text(a['excerpt'])}</p>
          </a>
"""

        recent_section = ""
        if recent:
            recent_section = f"""
    <section style="padding: var(--space-m) 0; border-top: 1px solid var(--color-line);">
      <div class="container">
        <p class="section-label reveal">Recent</p>
        <div class="journal-grid reveal">{recent_html}
        </div>
      </div>
    </section>
"""

        body = f"""
  <main>
    <section class="lead">
      <div class="container">
        <h1 class="display reveal">The Journal.</h1>
        <p class="lead__support reveal">Notes on architecture, design, and the practice. Written by Studio Dotbox.</p>
      </div>
    </section>

    <section style="padding: var(--space-m) 0; border-top: 1px solid var(--color-line);">
      <div class="container">
        <p class="section-label reveal">Featured</p>
        <a href="journal/{featured['slug']}.html" class="journal-featured reveal">
          <div class="journal-featured__image">
            <img src="images/journal/{featured['slug']}-thumb.jpg" alt="{normalise_text(featured['title'])}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Featured article image]';">
          </div>
          <div class="journal-featured__content">
            <p class="article-meta"><span>{category_label(featured['category'])}</span><span>{featured['reading_time']}</span><span>{featured['publish_date']}</span></p>
            <h2 class="journal-featured__title">{normalise_text(featured['title'])}</h2>
            <p style="font-size: var(--size-body-large); color: var(--color-text-muted); line-height: 1.5;">{normalise_text(featured['excerpt'])}</p>
            <span class="link-arrow">Read the article →</span>
          </div>
        </a>
      </div>
    </section>
{recent_section}
    <section style="padding: var(--space-m) 0; border-top: 1px solid var(--color-line);">
      <div class="container">
        <p class="section-label reveal">Browse the Journal by topic</p>
        <p class="reveal" style="font-size: var(--size-body-large); font-weight: var(--weight-light); color: var(--color-text-muted); line-height: 1.8;">Design thinking · Process · Residential · Retail · Commercial · Hospitality · Urban planning · Consultancy · Materials · Site notes · Studio diary</p>
      </div>
    </section>
  </main>
"""

    write_page("journal.html", "Journal | Studio Dotbox",
        "Notes on architecture, design, and the practice. Written by Studio Dotbox.",
        body)


def build_journal_article(article):
    a = article
    body_html = paragraphs_to_html(a['body'])

    body = f"""
  <main>
    <article>
      <header class="article-header">
        <div class="container-narrow">
          <p class="article-meta reveal">
            <span>By Pratyush Sharma</span>
            <span>{category_label(a['category'])}</span>
            <span>{a['reading_time']}</span>
            <span>{a['publish_date']}</span>
          </p>
          <h1 class="article-title reveal">{normalise_text(a['title'])}</h1>
        </div>
      </header>

      <div class="container-narrow">
        <div class="article-feature-image reveal">
          <img src="../images/journal/{a['slug']}.jpg" alt="{normalise_text(a['title'])}" onerror="this.style.display='none'; this.parentElement.innerHTML='[Article hero: 1800 × 1200 px, save as images/journal/{a['slug']}.jpg]';">
        </div>
      </div>

      <div class="article-body">
        <div class="container-narrow">
          <div class="reveal">
{body_html}
          </div>
        </div>
      </div>

      <footer class="article-signature">
        <div class="container-narrow">
          <p class="reveal">Written by Pratyush Sharma, Principal Architect, Studio Dotbox.</p>
          <p class="reveal">{a['publish_date']}</p>
        </div>
      </footer>
    </article>

    <section style="padding: var(--space-m) 0; border-top: 1px solid var(--color-line);">
      <div class="container-narrow" style="text-align: center;">
        <p class="reveal" style="font-size: var(--size-body-large); font-weight: var(--weight-light); color: var(--color-text-muted); margin-bottom: var(--space-s);">If this resonates, we would like to hear from you.</p>
        <a href="../contact.html" class="btn btn--primary reveal">Start the conversation →</a>
      </div>
    </section>

    <section class="continue">
      <div class="container">
        <p class="continue__label reveal">More from the Journal</p>
        <div class="continue__links reveal"><a href="../journal.html" class="link-arrow">All articles →</a></div>
      </div>
    </section>
  </main>
"""
    write_page(f"journal/{a['slug']}.html",
        f"{a['title']} | Studio Dotbox Journal",
        normalise_text(a['excerpt']) or normalise_text(a['title']),
        body, depth=1)


# ==================== MAIN ====================

def build_sitemap():
    """Emit sitemap.xml listing every built page as its canonical www https URL."""
    urls = sorted({canonical_url(f) for f in PAGES_BUILT})
    lines = ['<?xml version="1.0" encoding="UTF-8"?>',
             '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for u in urls:
        lines.append(f"  <url><loc>{u}</loc></url>")
    lines.append("</urlset>")
    (SITE_ROOT / "sitemap.xml").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Built: sitemap.xml ({len(urls)} urls)")


def build_robots():
    """Emit robots.txt allowing all crawlers and pointing to the sitemap."""
    content = f"User-agent: *\nAllow: /\n\nSitemap: {BASE_URL}/sitemap.xml\n"
    (SITE_ROOT / "robots.txt").write_text(content, encoding="utf-8")
    print("Built: robots.txt")


def main():
    print("Studio Dotbox: building all pages...")
    print(f"Reading content from: {EXCEL_FILE}")
    projects, articles = read_excel()
    print(f"Found {len(projects)} project(s) and {len(articles)} article(s).")
    print()

    # Featured items for Home page
    featured_projects = [p for p in projects if p['featured']] or projects[:5]
    featured_articles = [a for a in articles if a['featured']] or articles[:2]

    # Static pages
    build_home(featured_projects, featured_articles)
    build_studio()
    build_philosophy()
    build_services()
    build_how_we_work()
    build_contact()
    build_work_with_us()
    build_terms()
    build_privacy()
    build_404()

    # Excel-driven pages
    build_work_archive(projects)
    for project in projects:
        build_project_page(project)

    build_journal_archive(articles)
    for article in articles:
        build_journal_article(article)

    # SEO artifacts (must run last, after PAGES_BUILT is fully populated)
    build_sitemap()
    build_robots()

    print()
    print("Build complete.")
    print(f"Open index.html in your browser to preview the site.")


if __name__ == '__main__':
    main()
